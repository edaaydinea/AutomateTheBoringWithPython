import openpyxl, csv, os

# Ensure the output directory exists
os.makedirs('csvFiles', exist_ok=True)

# Loop through all files in the current working directory and subdirectories.
for foldername, subfolders, filenames in os.walk('./'):
    for filename in filenames:
        # Skip non-xlsx files and load the workbook object.
        if not filename.endswith('.xlsx'):
            continue
        else:
            excelFile = os.path.join(foldername, filename)
            wb = openpyxl.load_workbook(excelFile)
            # Loop through every sheet in the workbook.
            for sheetName in wb.sheetnames:
                sheet = wb[sheetName]

                # Create the CSV filename from the Excel filename and sheet title.
                excelFilenameWithoutExt = os.path.splitext(os.path.basename(excelFile))[0]
                csvFilename = f'{excelFilenameWithoutExt}_{sheetName}.csv'

                # Define the full path to save the CSV in the desired folder.
                csvFilePath = os.path.join('csvFiles', csvFilename)

                # Create the csv.writer object for this CSV file.
                with open(csvFilePath, 'w', newline='') as csvFile:
                    csvWriter = csv.writer(csvFile)

                    # Loop through every row in the sheet.
                    for rowNum in range(1, sheet.max_row + 1):
                        rowData = []  # Append each cell to this list.
                        
                        # Loop through each cell in the row.
                        for colNum in range(1, sheet.max_column + 1):
                            cellValue = sheet.cell(row=rowNum, column=colNum).value
                            rowData.append(cellValue)
                        
                        # Write the rowData list to the CSV file.
                        csvWriter.writerow(rowData)

                print(f'Created {csvFilePath}')
