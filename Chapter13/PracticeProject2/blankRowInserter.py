import sys
import openpyxl

# Get the arguments from the command line
if len(sys.argv) < 4:
    print("Usage: python blankRowInserter.py <N> <M> <filename>")
    sys.exit()

N = int(sys.argv[1])  # Starting row where blanks are inserted
M = int(sys.argv[2])  # Number of blank rows to insert
filename = sys.argv[3]  # The Excel file to modify

# Load the workbook and select the active sheet
wb = openpyxl.load_workbook(filename)
sheet = wb.active

# Get the number of rows in the current sheet
max_row = sheet.max_row

# Create a new workbook to store the modified spreadsheet
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# Copy the contents of the old spreadsheet to the new one with M blank rows after row N
for row in range(1, N):
    for col in range(1, sheet.max_column + 1):
        new_sheet.cell(row=row, column=col).value = sheet.cell(row=row, column=col).value

# Skip M rows starting from row N, and copy the remaining rows
for row in range(N, max_row + 1):
    for col in range(1, sheet.max_column + 1):
        new_sheet.cell(row=row + M, column=col).value = sheet.cell(row=row, column=col).value

# Save the new workbook to a file
new_filename = f'updated_{filename}'
new_wb.save(new_filename)

print(f"Inserted {M} blank rows at row {N} in '{filename}' and saved the result as '{new_filename}'")
