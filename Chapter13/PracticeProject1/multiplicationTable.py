import sys
import openpyxl
from openpyxl.styles import Font

# Get the number N from command line
if len(sys.argv) < 2:
    print("Usage: py multiplicationTable.py <N>")
    sys.exit()

N = int(sys.argv[1])

# Create a new workbook and get the active sheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = f'{N}x{N} Multiplication Table'

# Create a bold font for labels
bold_font = Font(bold=True)

# Fill the first row and column with labels and make them bold
for i in range(1, N + 1):
    # Row 1 and Column A are labels
    sheet.cell(row=1, column=i+1).value = i  # Top labels
    sheet.cell(row=1, column=i+1).font = bold_font  # Make them bold
    
    sheet.cell(row=i+1, column=1).value = i  # Side labels
    sheet.cell(row=i+1, column=1).font = bold_font  # Make them bold

# Fill in the multiplication table
for row in range(2, N + 2):
    for col in range(2, N + 2):
        row_label = sheet.cell(row=row, column=1).value  # Side label
        col_label = sheet.cell(row=1, column=col).value  # Top label
        sheet.cell(row=row, column=col).value = row_label * col_label  # Multiply row label by column label

# Save the workbook
wb.save(f'multiplication_table_{N}x{N}.xlsx')

print(f"Multiplication table of size {N}x{N} saved as 'multiplication_table_{N}x{N}.xlsx'")
