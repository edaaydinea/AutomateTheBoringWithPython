import openpyxl
import sys

# Load the workbook and select the active sheet
wb = openpyxl.load_workbook('input_spreadsheet.xlsx')
sheet = wb.active

# Iterate over each column in the spreadsheet
for col_num in range(1, sheet.max_column + 1):
    # Create a text file for each column
    with open(f'column_{col_num}.txt', 'w') as file:
        # Iterate over the rows in the column
        for row_num in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_num, column=col_num).value
            if cell_value:  # If the cell is not empty
                file.write(f"{cell_value}\n")

print("Saved each column into separate text files.")
