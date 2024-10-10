import openpyxl

# Load the workbook and select the active sheet
filename = 'original_file.xlsx'  # Change this to the filename you are working with
wb = openpyxl.load_workbook(filename)
sheet = wb.active

# Create a list of lists to hold the cell data
sheet_data = []

# Read the sheet into sheet_data
for row in sheet.iter_rows(values_only=True):
    sheet_data.append(list(row))

# Get the dimensions of the original sheet
max_row = len(sheet_data)
max_col = len(sheet_data[0]) if max_row > 0 else 0

# Create a new workbook for the transposed data
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# Write the transposed data to the new sheet
for row in range(max_row):
    for col in range(max_col):
        # Transpose the value at row, col to col, row
        new_sheet.cell(row=col + 1, column=row + 1).value = sheet_data[row][col]

# Save the transposed workbook
new_filename = 'inverted_' + filename
new_wb.save(new_filename)

print(f"Spreadsheet rows and columns have been inverted and saved as '{new_filename}'")
